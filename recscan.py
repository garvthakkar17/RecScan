import dns.resolver
import argparse
import sys
from termcolor import colored
import itertools
import openpyxl
from openpyxl.styles import Alignment

# Rainbow colors for ASCII art
rainbow_colors = itertools.cycle(["red", "yellow", "green", "blue", "magenta", "cyan"])

# Colorized ASCII Art for RecScan with rainbow colors
ascii_art = """
d8888b. d88888b  .o88b. .d8888.  .o88b.  .d8b.  d8b   db
88  `8D 88'     d8P  Y8 88'  YP d8P  Y8 d8' `8b 888o  88
88oobY' 88ooooo 8P      `8bo.   8P      88ooo88 88V8o 88
88`8b   88~~~~~ 8b        `Y8b. 8b      88~~~88 88 V8o88
88 `88. 88.     Y8b  d8 db   8D Y8b  d8 88   88 88  V888
88   YD Y88888P  `Y88P' `8888Y'  `Y88P' YP   YP VP   V8P
"""

# Function to print rainbow colored ASCII art
def print_rainbow_art(art):
    for line in art.splitlines():
        colored_line = "".join([colored(char, next(rainbow_colors)) for char in line])
        print(colored_line)

# Function to check specific DNS record
def check_record(domain, record_type):
    try:
        answers = dns.resolver.resolve(domain, record_type)
        return [answer.to_text() for answer in answers]
    except dns.resolver.NoAnswer:
        return None
    except dns.resolver.NXDOMAIN:
        return None
    except Exception as e:
        return [f"Error: {e}"]

# Function to analyze a domain and save the result to an Excel file
def analyze_domain(domain, ws, row):
    print(colored(f"\nAnalyzing domain: {domain}", "cyan"))
    print("-" * 50)

    records = {
        "TXT": check_record(domain, "TXT"),
        "MX": check_record(domain, "MX"),
        "NS": check_record(domain, "NS"),
        "A": check_record(domain, "A"),
        "CNAME": check_record(domain, "CNAME"),
        "AAAA": check_record(domain, "AAAA"),
        "SPF": check_record(domain, "TXT"),  # SPF is a TXT record
        "DKIM": check_record(f"_domainkey.{domain}", "TXT"),
        "DMARC": check_record(f"_dmarc.{domain}", "TXT"),
        "SOA": check_record(domain, "SOA"),
        "PTR": check_record(domain, "PTR"),
        "SRV": check_record(domain, "SRV"),
        "CAA": check_record(domain, "CAA")
    }

    total_rows = sum(len(record_list) if record_list else 1 for record_list in records.values())
    ws.merge_cells(start_row=row, start_column=1, end_row=row + total_rows - 1, end_column=1)
    ws.cell(row=row, column=1, value=domain).alignment = Alignment(horizontal='center', vertical='center')

    for record_type, record_list in records.items():
        if record_list:
            ws.merge_cells(start_row=row, start_column=2, end_row=row + len(record_list) - 1, end_column=2)
            ws.cell(row=row, column=2, value=f"{record_type} Record").alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=row, start_column=4, end_row=row + len(record_list) - 1, end_column=4)
            ws.cell(row=row, column=4, value="Yes").alignment = Alignment(horizontal='center', vertical='center')

            for record in record_list:
                ws.cell(row=row, column=3, value=record)
                row += 1
        else:
            ws.cell(row=row, column=2, value=f"{record_type} Record").alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=3, value="Missing")
            ws.cell(row=row, column=4, value="No").alignment = Alignment(horizontal='center', vertical='center')
            row += 1

    return row

# Function to create an Excel file
def create_excel_file(domains, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DNS Records"

    ws.cell(row=1, column=1, value="Domain")
    ws.cell(row=1, column=2, value="Record Type")
    ws.cell(row=1, column=3, value="Details")
    ws.cell(row=1, column=4, value="Found")

    row = 2
    for domain in domains:
        row = analyze_domain(domain, ws, row)

    ws.cell(row=row + 2, column=1, value="Developed by Garv Thakkar").alignment = Alignment(horizontal='center', vertical='center')
    wb.save(output_file)
    print(f"Results saved to {output_file}")

# Main function
def main():
    print_rainbow_art(ascii_art)
    print(colored("Welcome to RecScan - DNS Record Checker\n", "yellow"))
    print(colored("Developed by Garv Thakkar\n", "green"))

    parser = argparse.ArgumentParser(
        description="Check DNS records (SPF, DKIM, DMARC, etc.) for domains.",
        epilog="Example usage:\n"
               "  Single domain: recscan.py domain.com\n"
               "  List of domains: recscan.py -l domains.txt",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument(
        "domain",
        nargs="?",
        help="The domain to check (for a single domain)."
    )
    parser.add_argument(
        "-l", "--list",
        help="Path to a text file containing a list of domains (one per line)."
    )
    parser.add_argument(
        "-o", "--output",
        help="Path to save the results in an Excel file."
    )
    args = parser.parse_args()

    domains = []

    if args.domain:
        domains.append(args.domain)
    elif args.list:
        try:
            with open(args.list, "r") as file:
                domains = [line.strip() for line in file if line.strip()]
        except FileNotFoundError:
            print(colored(f"Error: File '{args.list}' not found.", "red"))
            sys.exit(1)
    else:
        print(colored("Error: No domain or list provided. Use -h for help.", "red"))
        sys.exit(1)

    if args.output:
        create_excel_file(domains, args.output)
    else:
        for domain in domains:
            print(f"\nResults for domain: {domain}")
            records = {
                "TXT": check_record(domain, "TXT"),
                "MX": check_record(domain, "MX"),
                "NS": check_record(domain, "NS"),
                "A": check_record(domain, "A"),
                "CNAME": check_record(domain, "CNAME"),
                "AAAA": check_record(domain, "AAAA"),
                "SPF": check_record(domain, "TXT"),
                "DKIM": check_record(f"_domainkey.{domain}", "TXT"),
                "DMARC": check_record(f"_dmarc.{domain}", "TXT"),
                "SOA": check_record(domain, "SOA"),
                "PTR": check_record(domain, "PTR"),
                "SRV": check_record(domain, "SRV"),
                "CAA": check_record(domain, "CAA")
            }

            for record_type, record_list in records.items():
                status = "\u2714 Found" if record_list else "\u2718 Not Found"
                color = "green" if record_list else "red"
                print(colored(f"{record_type} Record: {status}", color))
                if record_list:
                    for record in record_list:
                        print(colored(f"  - {record}", "yellow"))

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nExiting RecScan. Bye!")
        sys.exit(0)
